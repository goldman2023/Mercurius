import numpy as np
import random
import pandas as pd
import pymysql
import json
import colorsys
from datetime import timedelta, datetime
from xlsxwriter import *
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl.drawing.image import Image

json_file = open("db_config.json")
db_info = json.load(json_file)

db_ip_address = db_info['db_ip_address']
db_name = db_info['db_name']
print('db name: ', db_name)
print('ip address: ', db_ip_address)


class Database:
    # create database with name specifying the specific database to connect with
    def __init__(self, db_name, db_ip_address):
        self.db_name = db_name
        self.db_ip_address = db_ip_address

    # create and return a mysql connection object
    def create_db_connection(self):

        try:
            # connection for EC2 Test
            conn = pymysql.connect(host=self.db_ip_address, user="root", password="LNTnjFxYvBjDkPw5", database=self.db_name)
            # print('Successfully connected to database')
            return conn
        except Exception as error:
            print('connection failed!', error)
            exit()

    # retrieve all business rules and return as a python dictionary
    def get_all_business_rules(self):

        conn = self.create_db_connection()

        try:
            with conn.cursor() as cursor:
                sql_statement = "SELECT Parameter, Value, Shift_Length FROM ref_faa_business_rules"
                cursor.execute(sql_statement)
                business_rules = cursor.fetchall()

                business_rules_dict = {}
                for rule in business_rules:
                    try:
                        business_rules_dict[rule[2]][rule[0]] = rule[1]
                    except KeyError:
                        business_rules_dict[rule[2]] = {}
                        business_rules_dict[rule[2]][rule[0]] = rule[1]

                return business_rules_dict

        except Exception as error:
            print(error)
        finally:
            conn.close()

    # get the user_id of the user that created a given bid_schedule
    def get_user_id_of_bid_schdule(self, bid_schedule_id):
        conn = self.create_db_connection()

        try:
            with conn.cursor() as cursor:
                sql_statement = "SELECT bid_manager_id FROM bid_schedule WHERE bid_schedule_id = {}" \
                    .format(bid_schedule_id)
                cursor.execute(sql_statement)
                user_id = cursor.fetchone()[0]

                return user_id

        except Exception as error:
            print(error)
        finally:
            conn.close()
     # get the total round of  a given bid_schedule
    def get_total_round_of_bid_schdule(self, bid_schedule_id):
        conn = self.create_db_connection()

        try:
            with conn.cursor() as cursor:
                sql_statement = "SELECT bid_round_count FROM bid_schedule WHERE bid_schedule_id = {}" \
                    .format(bid_schedule_id)
                cursor.execute(sql_statement)
                bid_round_count = cursor.fetchone()[0]

                return bid_round_count

        except Exception as error:
            print(error)
        finally:
            conn.close()

    # get all shiftline_name, employee_id, schedule_name tuples for a particular bid_schedule_id
    def get_employee_info_and_shiftlines(self, bid_schedule_id):
        conn = self.create_db_connection()

        employee_info = {}
        employee_shifts = {}
        schedule_ids = []
        schedule_name_shiftline_sequence = {}
        bid_schedule_name_list = []
        shiftline_schedule_name_list = []

        try:
            with conn.cursor() as cursor:
                sql_statement = "select distinct ed.employee_id, ed.employee_operating_initials, ed.first_name, ed.last_name, ed.seniority_ranking, ed.vacation_leave, sbt.shift_line_schedule_id_ref " \
                                 ",case when sbt.emp_id_ref is null then null else sbt.shift_line_id_ref end as shift_line_id_ref " \
                                 ",sbt.shift_line_schedule_name " \
                                 ",case when sbt.emp_id_ref is null then null else sbt.shiftline_sequence end as shiftline_sequence " \
                                 ",case when sbt.emp_id_ref is null then null else sbt.shiftline_name end as shiftline_name " \
                                 ",sbt.bid_schedule_name " \
                                "from bidschedule_employeedetails_map bem left join employee_details ed " \
                                "on bem.emp_id_ref = ed.employee_id " \
                                "left join shiftline_bid_transaction_table sbt " \
                                "on sbt.emp_id_ref = bem.emp_id_ref " \
                                "and sbt.bid_schedule_id_ref = bem.bid_schedule_id_ref " \
                                "where bem.bid_schedule_id_ref = {} " \
                              .format(bid_schedule_id)
                cursor.execute(sql_statement)
                tuples = cursor.fetchall()
                sql_statement3 = "select wt.emp_id_ref, wt.shiftline_bid_status from window_transaction_table wt where bid_schedule_id_ref = {} and bid_round_seq = 1 " \
                    .format(bid_schedule_id)
                cursor.execute(sql_statement3)
                tuples_childs = cursor.fetchall()
                if len(tuples) == 0:
                    sql_statement2 = "SELECT ed.employee_id, ed.employee_operating_initials, ed.first_name, ed.last_name, " \
                                     "ed.seniority_ranking, ed.vacation_leave, sls.shift_line_schedule_id, sls.shift_line_schedule_name, bs.bid_schedule_name " \
                                     "FROM  bidschedule_employeedetails_map bem, employee_details ed, bidschedule_shiftlineschedule_map bsm, shift_line_schedule sls, bid_schedule bs " \
                                     "WHERE bem.bid_schedule_id_ref = {} AND bem.emp_id_ref = ed.employee_id AND bem.bid_schedule_id_ref = bsm.bid_schedule_id_ref AND bsm.shift_line_schedule_id_ref = sls.shift_line_schedule_id AND bem.bid_schedule_id_ref = bs.bid_schedule_id"\
                        .format(bid_schedule_id)
                    cursor.execute(sql_statement2)
                    tuples = cursor.fetchall()

                    for tuple in tuples:
                        (employee_id, initials, first_name, last_name, seniority_ranking, vacation_leave, schedule_id, shift_line_schedule_name, bid_schedule_name) = tuple
                        if employee_id not in employee_info.keys():
                            employee_info[employee_id] = {
                                "initials": initials,
                                "first_name": first_name,
                                "last_name": last_name,
                                "seniority_ranking": seniority_ranking,
                                "vacation_leave": vacation_leave
                            }
                        if schedule_id != None :
                            try:
                                employee_shifts[employee_id][schedule_id] = ''
                            except KeyError:
                                employee_shifts[employee_id] = {}
                                employee_shifts[employee_id][schedule_id] = ''

                            if schedule_id not in schedule_ids:
                                schedule_ids.append(schedule_id)

                        if shift_line_schedule_name != None:
                            if shift_line_schedule_name not in shiftline_schedule_name_list:
                                shiftline_schedule_name_list.append(shift_line_schedule_name)

                            try:
                                schedule_name_shiftline_sequence[employee_id][shift_line_schedule_name] = ''
                            except KeyError:
                                schedule_name_shiftline_sequence[employee_id] = {}
                                schedule_name_shiftline_sequence[employee_id][shift_line_schedule_name] = ''

                            if bid_schedule_name not in bid_schedule_name_list:
                                bid_schedule_name_list.append(bid_schedule_name)

                    for emp_id, sched_name_shift_line_seq in schedule_name_shiftline_sequence.items():
                        for sched_name in shiftline_schedule_name_list:
                            try:
                                x = schedule_name_shiftline_sequence[emp_id][sched_name]
                            except KeyError:
                                schedule_name_shiftline_sequence[emp_id][sched_name] = " "

                else:
                    for tuple in tuples:
                        (employee_id, initials, first_name, last_name, seniority_ranking, vacation_leave, schedule_id, shiftline_id, shift_line_schedule_name, shiftline_sequence, shiftline_name, bid_schedule_name) = tuple
                        for tuples_child in tuples_childs:
                            (emp_id_ref, shiftLine_bid_status)=tuples_child
                            if emp_id_ref==employee_id:
                                shiftLine_bidstatus=shiftLine_bid_status
                        if employee_id not in employee_info.keys():
                            employee_info[employee_id] = {
                                "initials": initials,
                                "first_name": first_name,
                                "last_name": last_name,
                                "seniority_ranking": seniority_ranking,
                                "vacation_leave": vacation_leave,
                                "shiftLine_bid_status":shiftLine_bidstatus
                            }
                        if schedule_id != None and shiftline_id != None:
                            try:
                                employee_shifts[employee_id][schedule_id] = shiftline_id
                            except KeyError:
                                employee_shifts[employee_id] = {}
                                employee_shifts[employee_id][schedule_id] = shiftline_id

                            if schedule_id not in schedule_ids:
                                schedule_ids.append(schedule_id)

                        if shift_line_schedule_name != None and shiftline_sequence != None:
                            if shift_line_schedule_name not in shiftline_schedule_name_list:
                                shiftline_schedule_name_list.append(shift_line_schedule_name)

                            try:
                                schedule_name_shiftline_sequence[employee_id][shift_line_schedule_name] = shiftline_name
                            except KeyError:
                                schedule_name_shiftline_sequence[employee_id] = {}
                                schedule_name_shiftline_sequence[employee_id][shift_line_schedule_name] = shiftline_name

                            if bid_schedule_name not in bid_schedule_name_list:
                                bid_schedule_name_list.append(bid_schedule_name)
                        else:
                            if shift_line_schedule_name not in shiftline_schedule_name_list:
                                shiftline_schedule_name_list.append(shift_line_schedule_name)
                                print(shift_line_schedule_name)
                            try:
                                schedule_name_shiftline_sequence[employee_id][shift_line_schedule_name] = ''
                            except KeyError:
                                schedule_name_shiftline_sequence[employee_id] = {}
                                schedule_name_shiftline_sequence[employee_id][shift_line_schedule_name] = ''

                            if bid_schedule_name not in bid_schedule_name_list:
                                bid_schedule_name_list.append(bid_schedule_name)
                    for emp_id, sched_name_shift_line_seq in schedule_name_shiftline_sequence.items():
                        for sched_name in shiftline_schedule_name_list:
                            try:
                                x = schedule_name_shiftline_sequence[emp_id][sched_name]
                            except KeyError:
                                schedule_name_shiftline_sequence[emp_id][sched_name] = " "

                return employee_info, employee_shifts, schedule_ids, schedule_name_shiftline_sequence, bid_schedule_name_list

        except Exception as error:
            print(error)
        finally:
            conn.close()

    # Get all shiftlines for a schedule
    def get_shifts_to_replace_shiftline_id(self, schedule_id):
        conn = self.create_db_connection()
        try:
            with conn.cursor() as cursor:
                sql_statement = "SELECT shift_line_id, sun_shift, mon_shift, tue_shift, wed_shift, thu_shift" \
                                "fri_shift, sat_shift FROM  shift_line " \
                                "WHERE sh_schedule_id_ref = {}" \
                    .format(schedule_id)
                cursor.execute(sql_statement)
                tuples = cursor.fetchall()

                shiftlines = {}

                for tuple in tuples:
                    shiftlines[tuple[0]] = {
                        "sun": tuple[1],
                        "mon": tuple[2],
                        "tue": tuple[3],
                        "wed": tuple[4],
                        "thu": tuple[5],
                        "fri": tuple[6],
                        "sat": tuple[7],
                    }

                return shiftlines

        except Exception as error:
            print(error)
        finally:
            conn.close()

    # get all employee_id, vacation_start_date, vacation_end_date tuples for a particular bid_schedule_id
    def get_employee_id_vacation_start_end_tuples(self, bid_schedule_id):
        conn = self.create_db_connection()

        try:
            with conn.cursor() as cursor:
                sql_statement = "SELECT emp_id_ref, vacation_start_date, vacation_end_date, bid_round_seq, vacation_selected_inhours, vacation_hours " \
                                "FROM  vacation_bid_transaction_table " \
                                "WHERE bid_schedule_id_ref = {}" \
                    .format(bid_schedule_id)
                cursor.execute(sql_statement)
                tuples = cursor.fetchall()

                return tuples

        except Exception as error:
            print(error)
        finally:
            conn.close()

    # Get shift line schedule id used by a given bid schedule id for a particular date
    def get_schedule_start_end_dates(self, bid_schedule_id):
        conn = self.create_db_connection()

        try:
            with conn.cursor() as cursor:
                sql_statement = "SELECT shift_line_schedule_id_ref, shiftline_schedule_start_date, " \
                                "shiftline_schedule_end_date " \
                                "FROM bidschedule_shiftlineschedule_map " \
                                "WHERE bid_schedule_id_ref = {}".format(bid_schedule_id)
                cursor.execute(sql_statement)
                tuples = cursor.fetchall()

                return tuples

        except Exception as error:
            print(error)
        finally:
            conn.close()


# Load all business rules from database
db = Database(db_name, db_ip_address)
business_rules = db.get_all_business_rules()


# transfer all business rules (created in this file) to a different file
def transfer_business_rules():
    return business_rules


# retrieve data regarding core vs ancilliary shifts
def get_core_ancilliary_shift_data():

    data = {
        'day': {},
        'eve': {},
        'mid': {}
    }

    for parameter, value in business_rules[8].items():
        for day in data.keys():
            if 'core' in parameter:
                if day in parameter:
                    data[day]['core'] = value
            elif 'ancilliary' in parameter:
                if day in parameter:
                    data[day]['ancilliary'] = value

    return data

# retrieve data for no shift window
def get_no_shift_window():
    no_shift_window = {}

    for parameter, value in business_rules[8].items():
        if "no_shift" in parameter:
            if "start_time" in parameter:
                no_shift_window['start_time'] = '0{}00'.format(value) if value < 10 else '{}00'.format(value)
            if "end_time" in parameter:
                no_shift_window['end_time'] = '0{}00'.format(value) if value < 10 else '{}00'.format(value)

    return no_shift_window


def st_to_mt(shift):
    if shift == "X":
        return "RDO"
    elif shift == "vl":
        return "vl"
    elif shift == "M":
        return "2300"
    elif shift == "N":
        return "2400"
    elif shift == "1":
        return "1300"
    elif shift == "3":
        return "1500"
    elif shift == "4":
        return "1600"
    elif len(shift) == 1:
        return "0" + shift + "00"
    elif len(shift) == 2:
        return shift + "00"
    else:
        return shift


# converts a value in military time to an equivalent integer/float
def mt_to_int(military_time):

    hours = military_time[0:2]
    minutes = military_time[2:]

    if hours[0] == 0:
        hours = hours[1]
    hours = int(hours)

    minutes = round(int(minutes)/60, 2)

    if minutes == 0:
        time = hours
    else:
        time = hours + minutes
    return time


def mt_to_mt_range(military_time, shift_length):
    if military_time == "RDO":
        return military_time
    else:
        hour_start = int(military_time[0:2])
        hour_end = (hour_start + 8) % 24
        if hour_end == 0:
            hour_end = 24
        minutes = military_time[2:]
        return "{}:{} - {}:{}".format(str(hour_start), minutes, str(hour_end), minutes)

# generate schedule leave summary
def generate_schedule_leave_summary(bid_schedule_id, start_year, start_month, start_day, end_year, end_month, end_day):
    # create datetime objects for start and end date
    start_date = datetime(start_year, start_month, start_day)
    end_date = datetime(end_year, end_month, end_day)

    # Get employee info and the shift_line_id they bid for each shift_line_schedule of this bid_schedule
    employee_info, employee_shifts, schedule_ids, schedule_name_shiftline_sequence, bid_schedule_name_list = db.get_employee_info_and_shiftlines(bid_schedule_id)
    total_bid_rounds=db.get_total_round_of_bid_schdule(bid_schedule_id)
    total_bid_rounds=int(total_bid_rounds)
    total_bid_rounds+=1
    print(total_bid_rounds)
    print(employee_info)
    print(employee_shifts)
    print(schedule_ids)
    print(schedule_name_shiftline_sequence)
    print(bid_schedule_name_list)

    shift_line_schedule_names = []

    for employee_id, schedule_name_shiftline_sequence_dict in schedule_name_shiftline_sequence.items():
        for schedule_name in schedule_name_shiftline_sequence_dict.keys():
            if schedule_name not in shift_line_schedule_names:
                if schedule_name!=None:
                    shift_line_schedule_names.append(schedule_name)

    print(shift_line_schedule_names)

    # Get the employee id and vacation start/end date they bid on for this bid schedule
    employee_vacation_bids = db.get_employee_id_vacation_start_end_tuples(bid_schedule_id)

    # Create lists of vacation days for each employee
    employee_vacation_bid_lists = {}
    for vacation_bid in employee_vacation_bids:
        (employee_id_vac, vac_start_date, vac_end_date, bid_round, picked_hours, shift_hours) = vacation_bid
        vac_days_delta = vac_end_date - vac_start_date
        # if(picked_hours/(vac_days_delta.days+1)==8):
        #     shiftline_hr=8
        # elif (picked_hours/(vac_days_delta.days+1)==10):
        #     shiftline_hr = 10
        # else:
        shiftline_hr = shift_hours
        for i in range(vac_days_delta.days + 1):
            try:
                employee_vacation_bid_lists[employee_id_vac][vac_start_date + timedelta(days=i)] = str(bid_round)+'-'+str(shiftline_hr)
            except KeyError:
                employee_vacation_bid_lists[employee_id_vac] = {}
                employee_vacation_bid_lists[employee_id_vac][vac_start_date + timedelta(days=i)] = str(bid_round)+'-'+str(shiftline_hr)

    print(employee_vacation_bid_lists)

    leave_summary = Workbook('schedule_leave_summary.xlsx', {'in_memory': True})

    worksheet1 = leave_summary.add_worksheet('Schedule Leave Summary')

    title_format = leave_summary.add_format({
        'font_size': 20,
        'bold': 1,
        'align': 'center',
        'bg_color':'#D3D3D3',
        'valign': 'vcenter'
    })

    header_format = leave_summary.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#D3D3D3',
        'border': 1,
        'text_wrap': True
    })
    header_data_format = leave_summary.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'text_wrap': True
    })
    total_picked_format = leave_summary.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True,
        'bg_color': '#FF4110',
        'border': 1,
    })

    worksheet1.merge_range('A1:T3', 'Schedule Leave Summary', title_format)

    worksheet1.insert_image(0, 0, 'image.png', {'image_width': 10, 'image_height': 10})

    worksheet1.merge_range('A4:C4', "Bid Schedule Name", header_format)
    if len(bid_schedule_name_list) > 0:
        worksheet1.merge_range('A5:C5', bid_schedule_name_list[0],header_data_format)
    else:
        worksheet1.merge_range('B5:D5', "Bid Schedule ID {}".format(bid_schedule_id),header_data_format)
    worksheet1.merge_range('D4:F4', "Bid Schedule Start Date", header_format)
    worksheet1.merge_range('D5:F5', "{}/{}/{}".format(start_month, start_day, start_year),header_data_format)
    worksheet1.merge_range('G4:I4', "Bid Schedule End Date", header_format)
    worksheet1.merge_range('G5:I5', "{}/{}/{}".format(end_month, end_day, end_year),header_data_format)
    worksheet1.merge_range('J4:P4', "Total Leave Accrued (days)"  , header_format)
    worksheet1.merge_range('J5:P5', "=sum(E10:E200)",header_data_format)
    worksheet1.merge_range('Q4:T4', "Total Leave Picked (Days)"  , header_format)
    # worksheet1.write("K5", "days")

    worksheet1.merge_range('D8:E8', "Accrued Leave", header_format)

    row = 7
    col = 6
    if len(shift_line_schedule_names) > 1:
        worksheet1.merge_range('{}:{}'.format(xl_rowcol_to_cell(row, col), xl_rowcol_to_cell(row, col + len(shift_line_schedule_names)-1)), "ShiftLine Schedule Names", header_format)
        col += len(shift_line_schedule_names)
    else:
        worksheet1.write(row, col, "ShiftLine Schedule Names", header_format)
        col += 1

    worksheet1.merge_range('{}:{}'.format(xl_rowcol_to_cell(row, col), xl_rowcol_to_cell(row, col + 1)), "Total Accrued Leave Picked", total_picked_format)

    row = 8
    col = 0
    colstyleCenter = leave_summary.add_format({
        'bold': False,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })
    colstyleLeft = leave_summary.add_format({
        'bold': False,
        'align': 'left',
        'valign': 'vcenter',

        'text_wrap': True
    })
    # worksheet1.write(row, col, 'Seniority',header_format)
    worksheet1.merge_range('A8:A9', "Seniority", header_format)
    worksheet1.set_column('A:A', 10,colstyleCenter)

    col += 1
    # worksheet1.write(row, col, 'Employee Name',header_format)
    worksheet1.merge_range('B8:B9', "Employee Name", header_format)
    worksheet1.set_column('B:B', 20,colstyleLeft)
    col += 1
    # worksheet1.write(row, col, 'Initials',header_format)
    worksheet1.merge_range('C8:C9', "Initials", header_format)
    worksheet1.set_column('C:C',10,colstyleCenter)
    col += 1
    worksheet1.write(row, col, 'Hours',header_format)
    col += 1
    worksheet1.write(row, col, 'Days',header_format)
    col += 1
    worksheet1.merge_range('F8:F9', "ShiftLine Bid Status", header_format)
    worksheet1.set_column('F:F', 20, colstyleLeft)
    col += 1

    if len(shift_line_schedule_names) > 0:
        for schedule in shift_line_schedule_names:
            worksheet1.write(row, col, schedule,header_format)

            col += 1
    else:
        worksheet1.write(row, col, "None",colstyleCenter)
        col += 1

    worksheet1.write(row, col, 'Hours', total_picked_format)
    col += 1
    worksheet1.write(row, col, 'Days', total_picked_format)
    worksheet1.merge_range('Q5:T5', "", header_data_format)
    worksheet1.write_formula(4,16, "=SUM({}:{})".format(xl_rowcol_to_cell(row+1, col),xl_rowcol_to_cell(200, col )),header_data_format)
    col += 1

    round_formats = []
    # round_formats.append(leave_summary.add_format({'bg_color': '#339966', 'border': 1, 'bold': True}))
    # round_formats.append(leave_summary.add_format({'bg_color': '#C1C100', 'border': 1, 'bold': True}))
    # round_formats.append(leave_summary.add_format({'bg_color': '#A05196', 'border': 1, 'bold': True}))
    #round_formats.append(leave_summary.add_format({'bg_color': '#34CCCC', 'border': 1, 'bold': True}))
    #round_formats.append(leave_summary.add_format({'bg_color': '#F5B084', 'border': 1, 'bold': True}))
   # round_formats.append(leave_summary.add_format({'bg_color': '#8EA9DC', 'border': 1, 'bold': True}))
    round_color_list=[]
    round_color_list_id = 1
    while round_color_list_id < total_bid_rounds:
        h, s, l = random.random(), 0.7 + random.random() / 2.0, 0.7 + random.random() / 6.0
        r, g, b = [int(256 * i) for i in colorsys.hls_to_rgb(h, l, s)]
        if r != g != b:
            if (int(r)>70 and int(g)>10 and int(b)>10 and int(r)<255 and int(g)<256 and int(b)<256):
                exists = '#%02x%02x%02x' % (r, g, b) in round_color_list
                if exists == False:
                    round_color_list.append('#%02x%02x%02x' % (r, g, b))
                    round_color_list_id += 1

    for round_color in round_color_list:
        round_formats.append(leave_summary.add_format({'bg_color': round_color,    'align': 'left','valign': 'vcenter', 'border': 1, 'bold': True}))
    for format in round_formats:
        format.set_rotation(90)

    worksheet1.set_column('{}:{}'.format(xl_rowcol_to_cell(row, col), xl_rowcol_to_cell(row, col + total_bid_rounds-2)), 3)

    for i in range(1, total_bid_rounds):
        worksheet1.merge_range("{}:{}".format(xl_rowcol_to_cell(row-1, col), xl_rowcol_to_cell(row, col)), 'Round {}'.format(i), round_formats[i-1])
        col += 1
    proxy_format= leave_summary.add_format({
        'bg_color': '#000000',
        'font_color': 'white',
        'bold':True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    proxy_format.set_rotation(90)
    worksheet1.set_column('{}:{}'.format(xl_rowcol_to_cell(row, col), xl_rowcol_to_cell(row, col)), 3)
    worksheet1.merge_range("{}:{}".format(xl_rowcol_to_cell(row - 1, col), xl_rowcol_to_cell(row, col)),
                           'Proxy',proxy_format)
    col += 1
    date_format = leave_summary.add_format({
        'bg_color': '#305496',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    dup_col=col
    # Create {day, schedule_id} dict for each date in bid_schedule
    num_days_delta = end_date - start_date
    all_dates = []
    leave_by_round_formats = []
    for round_color in round_color_list:
        leave_by_round_formats.append(leave_summary.add_format(
            {'bg_color': round_color, 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
    print(num_days_delta.days + 1)
    count=1
    dup_col+=1
    for i in range(num_days_delta.days + 1):
        if int(count)<int(num_days_delta.days + 1):
            for j in range(len(leave_by_round_formats)):
                if int(count) < int(num_days_delta.days + 1):
                    worksheet1.write(6, dup_col, 'R{}'.format(j+1),  leave_by_round_formats[j])
                    dup_col+=2
                    count+=2

    for i in range(num_days_delta.days + 1):
        date = start_date + timedelta(days=i)
        all_dates.append(date)
        row -= 1
        worksheet1.write(row, col, "{}/{}".format(date.strftime("%m"), date.strftime("%d")), date_format)
        row += 1
        worksheet1.write(row, col, date.strftime("%a"), date_format)
        col += 1

    row += 1
    col = 0
    emp_list_arr=[]
    for emp_id,emp_info_dict in employee_info.items() :
        emp_list_arr.append(emp_info_dict['seniority_ranking'])
    emp_list_arr.sort(reverse=False)
    for ranking in range(emp_list_arr[0], emp_list_arr[len(emp_list_arr)-1]+1):
        for emp_id, emp_info_dict in employee_info.items():
            if emp_info_dict['seniority_ranking'] == ranking:
                worksheet1.write(row, col, emp_info_dict['seniority_ranking'],colstyleCenter)
                col += 1
                worksheet1.write(row, col, "{} {}".format(emp_info_dict['first_name'], emp_info_dict['last_name']),colstyleLeft)
                col += 1
                worksheet1.write(row, col, emp_info_dict['initials'],colstyleCenter)
                col += 1
                worksheet1.write(row, col, int(emp_info_dict['vacation_leave']),colstyleCenter)
                col += 1
                worksheet1.write(row, col, int(emp_info_dict['vacation_leave'])/8,colstyleCenter)
                col += 1
                worksheet1.write(row, col,  emp_info_dict['shiftLine_bid_status'], colstyleLeft)
                col += 1
                for sched_name, sl_seq in schedule_name_shiftline_sequence[emp_id].items():
                    if sched_name!=None:
                        worksheet1.write(row, col, sl_seq,colstyleCenter)
                        col += 1
                total_hours_row = row
                total_hours_col = col
                # worksheet1.write_formula(row, col, "{}*8".format(xl_rowcol_to_cell(row, col+1)), total_picked_format)
                col += 1
                worksheet1.write_formula(row, col, "=SUM({}:{})".format(xl_rowcol_to_cell(row, col + 1), xl_rowcol_to_cell(row, col +total_bid_rounds-1)), total_picked_format)
                col += 1
                round_list=''
                for i in range(1, total_bid_rounds):
                    ten_hours = '{"R' + str(i) + '-10","R' + str(i) + '-9","'+'R'+str(i) + '-8"}'
                    worksheet1.write_formula(row, col, '=SUM(COUNTIF({}:{},{}))'.format(xl_rowcol_to_cell(row, col + total_bid_rounds - i), xl_rowcol_to_cell(row, col + total_bid_rounds - i + len(all_dates)),  ten_hours),colstyleCenter)
                    col += 1

                worksheet1.write_formula( row, col, '=COUNTIF({}: {}, "P")'.format(xl_rowcol_to_cell(row, col + total_bid_rounds - i), xl_rowcol_to_cell(row, col + total_bid_rounds - i + len(all_dates)), ''),colstyleCenter)
                col += 1
                for i in range(1, total_bid_rounds):
                    if(i==1):
                        round_list = round_list + 'COUNTIF(' + xl_rowcol_to_cell(row, col) + ':' + xl_rowcol_to_cell(row,col+ len(all_dates)) + ',"' + 'R' + str(i) + '-8")*8,COUNTIF(' + xl_rowcol_to_cell(row, col) + ':' + xl_rowcol_to_cell(row, col + len(all_dates)) + ',"' + 'R' + str(i) + '-9")*9,COUNTIF(' + xl_rowcol_to_cell(row, col) + ':' + xl_rowcol_to_cell(row,col + len(all_dates)) + ',"' + 'R' + str(i) + '-10")*10'
                    else:
                        round_list = round_list + ',COUNTIF(' + xl_rowcol_to_cell(row, col) + ':' + xl_rowcol_to_cell(row, col + len(all_dates)) + ',"' + 'R' + str(i) + '-8")*8,COUNTIF(' + xl_rowcol_to_cell(row, col) + ':' + xl_rowcol_to_cell(row, col + len(all_dates)) + ',"' + 'R' + str(i) + '-9")*9,COUNTIF(' + xl_rowcol_to_cell(row, col) + ':' + xl_rowcol_to_cell(row, col + len(all_dates)) + ',"' + 'R' + str(i) + '-10")*10'

                worksheet1.write_formula(total_hours_row, total_hours_col, '=SUM({})'.format(round_list), total_picked_format)

                leave_by_round_formats = []

                # leave_by_round_formats.append(leave_summary.add_format({'bg_color': '#339966', 'border': 1,  'align': 'center', 'valign': 'vcenter', 'bold': True}))
                # leave_by_round_formats.append(leave_summary.add_format({'bg_color': '#C1C100', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
                # leave_by_round_formats.append(leave_summary.add_format({'bg_color': '#A05196', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
                # leave_by_round_formats.append(leave_summary.add_format({'bg_color': '#34CCCC', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
                # leave_by_round_formats.append(leave_summary.add_format({'bg_color': '#F5B084', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
                # leave_by_round_formats.append(leave_summary.add_format({'bg_color': '#8EA9DC', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
                for round_color in round_color_list:
                                    leave_by_round_formats.append(leave_summary.add_format({'bg_color': round_color, 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True}))
                try:
                    vac_date_bid_round_dict_for_emp = employee_vacation_bid_lists[emp_id]
                    for date in all_dates:
                        try:
                            bid_round = vac_date_bid_round_dict_for_emp[date]
                            worksheet1.write(row, col, "R{}-{}".format(bid_round.split("-")[0],bid_round.split("-")[1]), leave_by_round_formats[int(bid_round.split("-")[0]) - 1])
                            col += 1
                        except KeyError:
                            col += 1
                except KeyError:
                    pass

                row += 1
                col = 0

    leave_summary.close()
    return leave_summary


# Generate Posted Watch Schedule for a given Bid Schedule between dates (inclusive)
def generate_posted_watch_schedule(bid_schedule_id, start_year, start_month, start_day, end_year, end_month, end_day):
    # create datetime objects for start and end date
    start_date = datetime(start_year, start_month, start_day)
    end_date = datetime(end_year, end_month, end_day)

    # get user_id of user that created the bid schedule
    # this will be used to look up user_defined shifts
    db.get_user_id_of_bid_schdule(bid_schedule_id)

    # Get employee info and the shift_line_id they bid for each shift_line_schedule of this bid_schedule
    employee_info, employee_shifts, schedule_ids, schedule_name_shiftline_sequence = db.get_employee_info_and_shiftlines(bid_schedule_id)
    # employee_info sample: 24: {'initials': 'WA', 'first_name': 'Michael', 'last_name': 'Ruple', 'vacation_leave': 208}
    # employee_shifts sample: 24: {34: 576, 52: 1188}
        # 24 is employee id
        # 34 is id of schedule
        # 576 is the id of the shift_line they bid on
    print(employee_info)
    print(employee_shifts)

    # In employee_shifts, replace shift_line_id with the actual shifts
    for schedule_id in schedule_ids:
        shift_lines = db.get_shifts_to_replace_shiftline_id(schedule_id)

        for emp_id, sch_id_sl_id_dict in employee_shifts.items():
            shift_line_id = sch_id_sl_id_dict[schedule_id]
            sch_id_sl_id_dict[schedule_id] = shift_lines[shift_line_id]

    print(employee_shifts)

    # Get the employee id and vacation start/end date they bid on for this bid schedule
    employee_vacation_bids = db.get_employee_id_vacation_start_end_tuples(bid_schedule_id)

    # Create lists of vacation days for each employee
    employee_vacation_bid_lists = {}
    for vacation_bid in employee_vacation_bids:
        (employee_id_vac, vac_start_date, vac_end_date, bid_round) = vacation_bid
        vac_days_delta = vac_end_date - vac_start_date
        for i in range(vac_days_delta.days + 1):
            try:
                employee_vacation_bid_lists[employee_id_vac].append(vac_start_date + timedelta(days=i))
            except KeyError:
                employee_vacation_bid_lists[employee_id_vac] = []
                employee_vacation_bid_lists[employee_id_vac].append(vac_start_date + timedelta(days=i))
    print(employee_vacation_bid_lists)

    # Get start and end dates for each schedule in bid_schedule
    schedule_date_tuples = db.get_schedule_start_end_dates(bid_schedule_id)
    print(schedule_date_tuples)

    # Create {day, schedule_id} dict for each date in bid_schedule
    num_days_delta = end_date - start_date
    all_dates = {}
    for i in range(num_days_delta.days + 1):
        date = start_date + timedelta(days=i)

        for tuple in schedule_date_tuples:
            sch_id, s_date, e_date = tuple
            if s_date <= date <= e_date:
                all_dates[date] = sch_id
    print(all_dates)

    # Create calendar workbook
    days = ['sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat']
    calendar = Workbook('test.xlsx')
    worksheet1 = calendar.add_worksheet('Schedule')
    worksheet2 = calendar.add_worksheet('Employee Names')
    red_format = calendar.add_format({'bg_color': 'red'})
    day_date_format = calendar.add_format({'bg_color': 'gray', 'font_color': 'black'})

    # Fill in worksheet of employee names
    emp_info_row = 0
    emp_info_column = 0
    for employee_info_tuple in employee_info.values():
        worksheet2.write(emp_info_row, emp_info_column, employee_info_tuple["initials"])
        emp_info_column = 1
        worksheet2.write(emp_info_row, emp_info_column, "{} {}".format(employee_info_tuple["first_name"], employee_info_tuple["last_name"]))
        emp_info_column = 0
        emp_info_row += 1

    # Fill in day headers at top of sheet
    row = 0
    column = 0
    for day in days:
        worksheet1.write(row, column, day, day_date_format)
        column += 1

    # fill calendar with dates according to date range
    current_row_number = 1
    start_row_number = 1
    max_row_number = 1

    # track max number of characters used to write data
    # this will be used to set column_widths
    max_column_width = 0

    # create a dict to track total shifts compared to those actually working (minus those on vl)
    workload_totals_for_day = {}
    actual_working_totals_for_day = {}
    all_shifts = []

    for given_date, sch_id_for_date in all_dates.items():
        day = given_date.strftime("%a").lower()
        column = days.index(day)

        worksheet1.write(current_row_number, column, given_date.strftime("%x"), day_date_format)
        current_row_number += 1

        shift_emp_ids_dict = {}
        vac_emp_ids_dict = {}
        workload_totals_for_day[given_date.strftime("%x")] = {}
        actual_working_totals_for_day[given_date.strftime("%x")] = {}

        # check each employee
        for employee_id in employee_shifts.keys():
            shift = employee_shifts[employee_id][sch_id_for_date][day]
            # determine if employee is on vacation
            # if so, add employee initials to list of those on vacation leave
            try:
                if given_date in employee_vacation_bid_lists[employee_id]:
                    try:
                        vac_emp_ids_dict[st_to_mt(shift)].append(employee_info[employee_id]['initials'])
                    except KeyError:
                        vac_emp_ids_dict[st_to_mt(shift)] = []
                        vac_emp_ids_dict[st_to_mt(shift)].append(employee_info[employee_id]['initials'])
                # if not on vacation, add employee initials to list of their shift
                else:
                    try:
                        shift_emp_ids_dict[st_to_mt(shift)].append(employee_info[employee_id]['initials'])
                    except KeyError:
                        shift_emp_ids_dict[st_to_mt(shift)] = []
                        shift_emp_ids_dict[st_to_mt(shift)].append(employee_info[employee_id]['initials'])
            # if employee has no vacation leave, a key error will result
            # add employee shift
            except KeyError:
                shift = employee_shifts[employee_id][sch_id_for_date][day]
                try:
                    shift_emp_ids_dict[st_to_mt(shift)].append(employee_info[employee_id]['initials'])
                except KeyError:
                    shift_emp_ids_dict[st_to_mt(shift)] = []
                    shift_emp_ids_dict[st_to_mt(shift)].append(employee_info[employee_id]['initials'])

        # sort shifts before exporting to excel
        # fill workload and actual_working_totals for this day
        sorted_shift_emp_ids_dict = {}
        for sorted_shift in sorted(shift_emp_ids_dict.keys()):
            if sorted_shift not in all_shifts:
                all_shifts.append(sorted_shift)
            sorted_shift_emp_ids_dict[sorted_shift] = shift_emp_ids_dict[sorted_shift]
            workload_totals_for_day[given_date.strftime("%x")][sorted_shift] = len(shift_emp_ids_dict[sorted_shift])
            actual_working_totals_for_day[given_date.strftime("%x")][sorted_shift] = len(shift_emp_ids_dict[sorted_shift])

        sorted_vac_emp_ids_dict = {}
        for sorted_vac_shift in sorted(vac_emp_ids_dict.keys()):
            if sorted_vac_shift not in all_shifts:
                all_shifts.append(sorted_vac_shift)
            sorted_vac_emp_ids_dict[sorted_vac_shift] = vac_emp_ids_dict[sorted_vac_shift]
            try:
                workload_totals_for_day[given_date.strftime("%x")][sorted_vac_shift] += len(vac_emp_ids_dict[sorted_vac_shift])
            except KeyError:
                workload_totals_for_day[given_date.strftime("%x")][sorted_vac_shift] = len(vac_emp_ids_dict[sorted_vac_shift])

        # export shifts to excel
        for shift, emp_ids_list in sorted_shift_emp_ids_dict.items():
            shift_ids = shift + ": (" + str(len(emp_ids_list)) + ") - "
            for id in emp_ids_list:
                shift_ids += id + " "
            worksheet1.write(current_row_number, column, shift_ids)
            if len(shift_ids) > max_column_width:
                max_column_width = len(shift_ids)
            current_row_number += 1

        total_on_vl = 0
        vac_emp_id_string = ""
        for shift, emp_ids_list in sorted_vac_emp_ids_dict.items():
            for id in emp_ids_list:
                vac_emp_id_string += id + " "
                total_on_vl += 1
        if total_on_vl > 0:
            vac_emp_id_string = "vl: (" + str(total_on_vl) + ") - " + vac_emp_id_string
            worksheet1.write(current_row_number, column, vac_emp_id_string, red_format)
            current_row_number += 1

        if current_row_number > max_row_number:
            max_row_number = current_row_number

        current_row_number = start_row_number
        if day == "sat":
            current_row_number = max_row_number + 1
            start_row_number = max_row_number + 1
    worksheet1.set_column(0, 6, max_column_width * 1.1)

    # if 7 day calendar - output week totals to new sheet
    if len(workload_totals_for_day.keys()) == 7:
        all_shifts = sorted(all_shifts)
        worksheet3 = calendar.add_worksheet('Totals')
        workload_week_totals = {}
        actual_working_week_totals = {}

        totals_row = 0
        totals_column = 0
        worksheet3.write(totals_row, totals_column, "Shifts")
        totals_row += 1

        for shift in all_shifts:
            worksheet3.write(totals_row, totals_column, shift)
            totals_row += 1
            workload_week_totals[shift] = 0
            actual_working_week_totals[shift] = 0

        totals_row = 0
        totals_column += 1

        for day in workload_totals_for_day.keys():
            worksheet3.write(totals_row, totals_column, day)
            totals_row += 1

            for shift in all_shifts:
                try:
                    work_load = workload_totals_for_day[day][shift]
                    workload_week_totals[shift] += work_load
                except KeyError:
                    work_load = 0
                try:
                    actual_working = actual_working_totals_for_day[day][shift]
                    actual_working_week_totals[shift] += actual_working
                except KeyError:
                    actual_working = 0
                if actual_working < work_load:
                    worksheet3.write(totals_row, totals_column, "{}/{}".format(actual_working, work_load), red_format)
                else:
                    worksheet3.write(totals_row, totals_column, "{}/{}".format(actual_working, work_load))
                totals_row += 1
            totals_row = 0
            totals_column += 1

        worksheet3.write(totals_row, totals_column, "Totals")
        totals_row += 1

        for shift in all_shifts:
            if actual_working_week_totals[shift] < workload_week_totals[shift]:
                worksheet3.write(totals_row, totals_column, "{}/{}".format(actual_working_week_totals[shift], workload_week_totals[shift]), red_format)
            else:
                worksheet3.write(totals_row, totals_column, "{}/{}".format(actual_working_week_totals[shift], workload_week_totals[shift]))
            totals_row += 1
    calendar.close()


def generate_single_day_watch_schedule(bid_schedule_id, year, month, day):
    # create datetime objects for start and end date
    date = datetime(year, month, day)

    # Get employee info and the shift_line_id they bid for each shift_line_schedule of this bid_schedule
    employee_info, employee_shifts, schedule_ids, schedule_name_shiftline_sequence = db.get_employee_info_and_shiftlines(bid_schedule_id)
    # employee_info sample: 24: {'initials': 'WA', 'first_name': 'Michael', 'last_name': 'Ruple', 'vacation_leave': 208}
    # employee_shifts sample: 24: {34: 576, 52: 1188}
        # 24 is employee id
        # 34 is id of schedule
        # 576 is the id of the shift_line they bid on

    # In employee_shifts, replace shift_line_id with the actual shifts
    for schedule_id in schedule_ids:
        shift_lines = db.get_shifts_to_replace_shiftline_id(schedule_id)

        for emp_id, sch_id_sl_id_dict in employee_shifts.items():
            shift_line_id = sch_id_sl_id_dict[schedule_id]
            sch_id_sl_id_dict[schedule_id] = shift_lines[shift_line_id]

    print(employee_shifts)

    # Get the employee id and vacation start/end date they bid on for this bid schedule
    employee_vacation_bids = db.get_employee_id_vacation_start_end_tuples(bid_schedule_id)

    # Create lists of vacation days for each employee
    employee_vacation_bid_lists = {}
    for vacation_bid in employee_vacation_bids:
        (employee_id_vac, vac_start_date, vac_end_date, bid_round) = vacation_bid
        vac_days_delta = vac_end_date - vac_start_date
        for i in range(vac_days_delta.days + 1):
            try:
                employee_vacation_bid_lists[employee_id_vac].append(vac_start_date + timedelta(days=i))
            except KeyError:
                employee_vacation_bid_lists[employee_id_vac] = []
                employee_vacation_bid_lists[employee_id_vac].append(vac_start_date + timedelta(days=i))
    print(employee_vacation_bid_lists)

    # create {emp_id: shift} dict for this date
    shift_emp_id_dict = {}
    vac_emp_id_dict = {}

    schedule_id = schedule_ids[0]
    date_string = date.strftime("%a").lower()
    all_shifts = []

    for e_id in employee_shifts.keys():
        shift = st_to_mt(employee_shifts[e_id][schedule_id][date_string])
        try:
            if date in employee_vacation_bid_lists[e_id]:
                try:
                    vac_emp_id_dict[shift].append(employee_info[e_id]['initials'])
                except KeyError:
                    vac_emp_id_dict[shift] = []
                    vac_emp_id_dict[shift].append(employee_info[e_id]['initials'])
            else:
                try:
                    shift_emp_id_dict[shift].append(employee_info[e_id]['initials'])
                except KeyError:
                    shift_emp_id_dict[shift] = []
                    shift_emp_id_dict[shift].append(employee_info[e_id]['initials'])
            if shift not in all_shifts:
                all_shifts.append(shift)
        except KeyError:
            shift = st_to_mt(employee_shifts[e_id][schedule_id][date_string])
            try:
                shift_emp_id_dict[shift].append(employee_info[e_id]['initials'])
            except KeyError:
                shift_emp_id_dict[shift] = []
                shift_emp_id_dict[shift].append(employee_info[e_id]['initials'])
            if shift not in all_shifts:
                all_shifts.append(shift)

    # sort dicts
    sorted_shift_emp_id_dict = {}
    for sorted_shift in sorted(shift_emp_id_dict.keys()):
        sorted_shift_emp_id_dict[sorted_shift] = shift_emp_id_dict[sorted_shift]

    sorted_vac_emp_id_dict = {}
    for sorted_vac_shift in sorted(vac_emp_id_dict.keys()):
        sorted_vac_emp_id_dict[sorted_vac_shift] = vac_emp_id_dict[sorted_vac_shift]

    print(sorted_shift_emp_id_dict)
    print(sorted_vac_emp_id_dict)

    # Create calendar workbook
    calendar = Workbook('single_day.xlsx')
    worksheet1 = calendar.add_worksheet('Schedule')
    red_format = calendar.add_format({'bg_color': 'red'})
    day_date_format = calendar.add_format({'bg_color': 'gray', 'font_color': 'black'})

    # initialize rows and columns
    # put complete date at top of file
    row = 0
    column = 0
    worksheet1.write(row, column, "{}, {} {}, {}".format(date.strftime("%A"), date.strftime("%B"), date.strftime("%d"), date.strftime("%Y")))
    row = 2

    for shift, emp_id_list in sorted_shift_emp_id_dict.items():
        worksheet1.write(row, column, mt_to_mt_range(shift, 8))
        row += 1
        for id in emp_id_list:
            worksheet1.write(row, column, id)
            row += 1
        column += 1
        row = 2

    worksheet1.write(row, column, "Leave")
    row += 1

    for shift, vac_id_list in sorted_vac_emp_id_dict.items():
        for id in vac_id_list:
            worksheet1.write(row, column, "{} - {}".format(id, mt_to_mt_range(shift, 8)))
            row += 1

    worksheet1.set_column(0, 10, 15)
    calendar.close()

# generate_posted_watch_schedule(331, 2023, 5, 7, 2023, 5, 13)
# generate_single_day_watch_schedule(331, 2023, 5, 13)


class ImpossibleRdoError(Exception):
    pass


class CascadeError(Exception):
    pass


class ShiftAssignError(Exception):
    pass


def errors_in_shift_line(shift_line, shift_length):

    global business_rules
    business_rules_for_shift_length = business_rules[shift_length]
    number_of_allowed_consecutive_mid_shifts = business_rules_for_shift_length['number_of_consecutive_mid_shifts']

    business_rule_violation = False

    # track number of RDO and if consecutive RDO found
    num_of_rdo = 0
    consecutive_rdo_found = False

    # determine if any shifts violate sufficient rest rules
    for i in range(len(shift_line)):
        prev_shift = shift_line[i % 7]
        next_shift = shift_line[(i + 1) % 7]
        if prev_shift == "X":
            num_of_rdo += 1
            if next_shift == "X":
                consecutive_rdo_found = True
        if sufficient_rest_between_shifts(prev_shift, next_shift, shift_length):
            continue
        else:
            business_rule_violation = True
            break

    # check if correct number of RDO exist for given shift_length
    if shift_length == 8:
        if num_of_rdo != 2:
            business_rule_violation = True
            print("incorrect number of RDO for shift length of 8")

    if shift_length == 10:
        if num_of_rdo != 3:
            business_rule_violation = True
            print("incorrect number of RDO for shift length of 10")

    # check if a pair of consecutive RDO were found
    if not consecutive_rdo_found:
        business_rule_violation = True
        print("no consecutive RDO detected")

    # if no rest violation, check consecutive mid shifts

    # track if we are dealing with an exception case
    exception_case = True

    # look at each shift in order
    for i in range(len(shift_line)):
        consecutive_mid_count = 0
        consecutive_mid_index = i
        current_shift = shift_line[consecutive_mid_index]
        # keep counting next day until shift of type not encountered
        while determine_type_of_shift(current_shift, shift_length) == 'MID':
            # during while loop, if we encounter one non-exception shift, exception case does not apply
            if current_shift != 21:
                exception_case = False
            consecutive_mid_count += 1
            # check if maximum number of consecutive shift exceeded while not being an exception case
            if consecutive_mid_count > number_of_allowed_consecutive_mid_shifts and not exception_case:
                business_rule_violation = True
                break
            consecutive_mid_index = (consecutive_mid_index + 1) % 7
            current_shift = shift_line[consecutive_mid_index]
        if business_rule_violation:
            break

    if business_rule_violation:
        return True
    else:
        return False


# check if PSO follows business rules
# if not, return the first pair of shifts causing a problem
def errors_in_PSO(PSO, PSO_MT, RDO_is_contiguous, shift_length):

    global business_rules
    business_rules_for_shift_length = business_rules[shift_length]
    number_of_allowed_consecutive_mid_shifts = business_rules_for_shift_length['number_of_consecutive_mid_shifts']

    if not RDO_is_contiguous:
        if not sufficient_rest_between_shifts(PSO[0], PSO[1], shift_length):
            return [PSO_MT[0], PSO_MT[1]]
        if not sufficient_rest_between_shifts(PSO[2], PSO[3], shift_length):
            return [PSO_MT[2], PSO_MT[3]]
    else:
        for i in range(len(PSO)-1):
            prev_shift = PSO[i]
            next_shift = PSO[i + 1]
            if sufficient_rest_between_shifts(prev_shift, next_shift, shift_length):
                continue
            else:
                return 'Error: The selected work pattern shift {} followed by {} will violate business rules; please select valid shifts'.format(PSO_MT[i], PSO_MT[i + 1])

        exception_case = True
        consecutive_mid_counter = 0

        for i in range(len(PSO)):
            shift = PSO[i]

            if determine_type_of_shift(shift, shift_length) == "MID":
                if shift != 21:
                    exception_case = False
                consecutive_mid_counter += 1
                if consecutive_mid_counter > number_of_allowed_consecutive_mid_shifts and not exception_case:
                    return 'Error: Number of allowed consecutive mid-shifts exceeded'
            else:
                consecutive_mid_counter = 0

    # if no return statement is encountered above there is no error
    return False


# Helper function for outlier function below
def split_data_set_into_halves(data_set):
    data_set_length = len(data_set)

    # Case 1 - even length data set
    if data_set_length % 2 == 0:
        lower_data_half = data_set[0: int(data_set_length/2)]
        upper_data_half = data_set[int(data_set_length/2):]
    # Case 2 - odd length data set
    else:
        lower_data_half = data_set[0: int(data_set_length / 2)]
        upper_data_half = data_set[int(data_set_length / 2) + 1:]

    return lower_data_half, upper_data_half


# Helper function for outlier function below
def find_median(data_set):
    data_set_length = len(data_set)

    # Case 1 - even length data set
    if data_set_length % 2 == 0:
        median = (data_set[int(data_set_length / 2) - 1] + data_set[int(data_set_length / 2)]) / 2
    # Case 2 - odd length data set
    else:
        median = data_set[int(data_set_length / 2)]

    return median


# Identifies any shifts that have an unusually high quantity using an outlier test
def check_for_large_outliers(daily_shifts):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    quantities_data_set = []

    for day in days:
        for quantity in daily_shifts[day].values():
            quantities_data_set.append(quantity)

    quantities_data_set.sort()

    # split data set into upper and lower half
    lower_data_half, upper_data_half = split_data_set_into_halves(quantities_data_set)

    # Find Q1 and Q3
    Q1 = find_median(lower_data_half)
    Q3 = find_median(upper_data_half)

    # Find IQR
    IQR = Q3 - Q1

    # Define upper limit of acceptable data
    upper_bound = Q3 + (1.5 * IQR)

    # check for and store any quantities that are outlier
    outlier_located = False
    outliers = {}
    for day in days:
        for shift, quantity in daily_shifts[day].items():
            if quantity > upper_bound:
                outlier_located = True
                if shift[0] == "0":
                    shift = shift[1:]
                try:
                    outliers[shift][day] = quantity
                except KeyError:
                    outliers[shift] = {}
                    outliers[shift][day] = quantity

    if outlier_located:
        return outliers
    else:
        return False


# ensure daily shifts is in the correct order
def sort_daily_shifts(daily_shifts):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    sorted_daily_shifts = {}
    for day in days:
        sorted_daily_shifts[day] = {}
        for i in np.linspace(23, 47, 2401):
            if i % 1 == 0:
                i = int(i % 24)
            else:
                i = round(round(i, 2) % 24, 2)
            try:
                sorted_daily_shifts[day][i] = daily_shifts[day][i]
            except KeyError:
                continue
    return sorted_daily_shifts


# receives starting time of a shift in the range 0 - 23 and returns type of shift
def determine_type_of_shift(shift_start, shift_length):
    if shift_start == 'X':
        return 'rdo'
    elif shift_start == None or isinstance(shift_start, list):
        return 'empty'
    elif (7 - shift_length / 2) < shift_start <= (7 + shift_length / 2):
        return "DAY"
    elif (15 - shift_length / 2) < shift_start <= (15 + shift_length / 2):
        return "EVE"
    else:
        return "MID"


# Identify day with least shifts
# Used to identify day that should be assigned extra shifts
def identify_day_with_least_shifts(total_shifts_per_day):
    min_daily_shifts = 999999999

    for day, num_of_shifts in total_shifts_per_day.items():
        if num_of_shifts < min_daily_shifts:
            min_daily_shifts = num_of_shifts
            day_with_least_shifts = day

    return day_with_least_shifts


# new random shift generator
# Identify day with least number of shifts
# Assign shifts to that day
def generate_targeted_random_shifts_to_add(shifts_to_add, total_shifts_per_day, daily_shifts):

    # Track the number of potential shifts added on each day
    potential_shifts_added = {}

    # Select existing shifts that are not mid-shifts as potential
    for day, shifts in daily_shifts.items():
        potential_shifts_added[day] = {}
        for shift, quantity in shifts.items():
            if quantity == 0 or 22 <= shift < 24 or 0 <= shift < 6:
                continue
            else:
                potential_shifts_added[day][shift] = 0

    while shifts_to_add > 0:
        # identify day with least total shifts
        day_to_add_shift = identify_day_with_least_shifts(total_shifts_per_day)

        # Identify shift on day with fewest number of shifts added
        min_number_of_shifts = 9999999999
        for shift, quantity in potential_shifts_added[day_to_add_shift].items():
            if quantity < min_number_of_shifts:
                min_number_of_shifts = quantity
                shift_to_add = shift

        # Add shift to daily shifts
        daily_shifts[day_to_add_shift][shift_to_add] += 1
        # Increase total_shifts on day
        total_shifts_per_day[day_to_add_shift] += 1
        # Record that shift has been added
        potential_shifts_added[day_to_add_shift][shift_to_add] += 1

        # decrement number of shifts to add
        shifts_to_add -= 1

    return daily_shifts, total_shifts_per_day


def identify_busiest_shift_on_day(daily_shifts, day, potential_shifts):

    max_num_of_shifts = 0
    busiest_shift = None

    for shift, quantity in daily_shifts[day].items():
        if shift in potential_shifts:
            if quantity > max_num_of_shifts:
                potential_shifts.remove(shift)
                max_num_of_shifts = quantity
                busiest_shift = shift

    return busiest_shift, potential_shifts


def generate_busiest_workday_shifts_to_add(shifts_to_add, daily_shifts):
    # only add random shifts to weekdays
    days = ["MON", "TUE", "WED", "THU", "FRI"]

    # select all week day shifts as possible shifts to add
    potential_shifts_to_add = {}
    for day in days:
        potential_shifts_to_add[day] = []
        for shift in daily_shifts[day].keys():
            if 22 <= shift < 24 or 0 <= shift < 6:
                continue
            elif daily_shifts[day][shift] == 0:
                continue
            else:
                potential_shifts_to_add[day].append(shift)

    while shifts_to_add > 0:
        # if every day has been selected, reset the days list
        if len(days) == 0:
            days = ["MON", "TUE", "WED", "THU", "FRI"]

        # select day / remove from days list
        random_day = days.pop(random.randint(0, len(days)-1))

        # refill a day with potential shifts if every potential shift for a day has already been assigned
        if len(potential_shifts_to_add[random_day]) == 0:
            for shift in daily_shifts[random_day].keys():
                if 22 <= shift < 24 or 0 <= shift < 6:
                    continue
                elif daily_shifts[random_day][shift] == 0:
                    continue
                else:
                    potential_shifts_to_add[random_day].append(shift)

        # pick a random shift and remove from list for the random day
        random_shift, potential_shifts_to_add[random_day] = identify_busiest_shift_on_day(daily_shifts, random_day, potential_shifts_to_add[random_day])
        daily_shifts[random_day][random_shift] += 1

        shifts_to_add -= 1

    return daily_shifts


# calculate how many of each rdo pair/triple is needed based on length of rdo and numbers of total days off each day
def calc_num_of_each_rdo_sequence(num_of_days_in_rdo, RDO_pattern, RDO_is_contiguous, days_off_per_day, num_RDO_failures):

    # create coefficient matrix based on number of days in RDO
    coefficient_matrix = np.zeros((7, 7), dtype=int)

    for i in range(0, 7):
        for j in range(0, 7):
            coefficient_matrix[i, (i+j) % 7] = RDO_pattern[j]

    # find inverse of matrix (necessary to solve)
    inverse_coefficient_matrix = np.linalg.inv(coefficient_matrix)

    # Create matrix with total number of RDO for each day
    RDO_totals_matrix = np.zeros((7, 1), dtype=int)
    index = 0
    for day, quantity in days_off_per_day.items():
        RDO_totals_matrix[index, 0] = quantity
        index += 1

    # solve for the number of each RDO sequence we need
    RDO_sequence_matrix = np.matmul(inverse_coefficient_matrix, RDO_totals_matrix)

    if num_RDO_failures == 0:
        RDO_minimum = 1
    else:
        RDO_minimum = 0

    if num_of_days_in_rdo == 2:
        RDO_pair_triple_name_list = ["SAT_SUN", "SUN_MON", "MON_TUE", "TUE_WED", "WED_THU", "THU_FRI", "FRI_SAT"]
    elif num_of_days_in_rdo == 3:
        if RDO_is_contiguous:
            RDO_pair_triple_name_list = ["FRI_SAT_SUN", "SAT_SUN_MON", "SUN_MON_TUE", "MON_TUE_WED", "TUE_WED_THU", "WED_THU_FRI", "THU_FRI_SAT"]
        else:
            RDO_pair_triple_name_list = ["SAT_SUN_WED", "SUN_MON_THU", "MON_TUE_FRI", "TUE_WED_SAT", "WED_THU_SUN", "THU_FRI_MON", "FRI_SAT_TUE"]

    # create dictionary of the number of each RDO sequence
    RDO = {}
    for i in range(0, 7):
        if int(round(RDO_sequence_matrix[i][0], 0)) >= RDO_minimum:
            RDO[RDO_pair_triple_name_list[i]] = int(round(RDO_sequence_matrix[i][0], 0))
        else:
            raise ImpossibleRdoError

    return RDO


# determines number of shifts of a current type that already exist on a given day
def num_of_shifts_of_type_assigned_for_day(day, shift_type, df):
    counter = 0
    current_shifts_on_day = df[day]
    for shift in current_shifts_on_day:
        if shift == shift_type:
            counter += 1
    return counter


def list_of_potential_shifts(day, shifts_assigned_on, daily_shifts):
    potential_shift_list = []

    # Look at shift requirements passed as input
    for shift, quantity in daily_shifts[day].items():
        # Ignore any shift that has a quantity of 0 (does not need to be assigned)
        if quantity != 0:
            # Assign shift if number assigned is less than number required
            if shifts_assigned_on[day][shift] < quantity:
                potential_shift_list.append(shift)

    return potential_shift_list


# determine if the amount of time between two shifts is sufficient
def sufficient_rest_between_shifts(prev_shift, next_shift, shift_length):
    global business_rules

    # use 10 hour business rules for 9 hour shifts
    if shift_length == 9:
        shift_length_for_business_rules = 10
    else:
        shift_length_for_business_rules = shift_length

    business_rules_for_given_shift_length = business_rules[shift_length_for_business_rules]

    if prev_shift == 'X' or next_shift == 'X':
        return True

    if isinstance(prev_shift, list) or isinstance(next_shift, list):
        return True

    # Case 1: prev_shift starts on given day
    if 0 <= prev_shift <= (23 - shift_length / 2):
        # Sub-Case 1: next_shift starts on given day
        if 0 <= next_shift <= (23 - shift_length / 2):
            if prev_shift < next_shift:
                return True
            else:
                time_between_shift_starts = (24 - prev_shift) + next_shift

        # Sub-Case 2: next_shift starts on previous day
        elif (23 - shift_length / 2) < next_shift < 24:
            time_between_shift_starts = next_shift - prev_shift

    # Case 2: prev_shift starts on previous day
    elif (23 - shift_length / 2) < prev_shift < 24:
        # Sub-Case 1: next_shift starts on given day
        if 0 <= next_shift <= (23 - shift_length / 2):
            return True

        # Sub-Case 2: next_shift starts on previous day
        elif (23 - shift_length / 2) < next_shift < 24:
            time_between_shift_starts = (24 - prev_shift) + next_shift

    rest = round(time_between_shift_starts - shift_length, 2)

    prev_shift_type = determine_type_of_shift(prev_shift, shift_length)
    next_shift_type = determine_type_of_shift(next_shift, shift_length)

    sufficient_rest = True

    # check if rest is sufficient for varying cases
    if prev_shift_type == "DAY" and next_shift_type == "MID":
        if shift_length == 8 and prev_shift <= 5.5:
            sufficient_rest = False
        else:
            if rest < business_rules_for_given_shift_length['time_between_day_shift_to_mid_shift']:
                sufficient_rest = False
    elif prev_shift_type == "EVE" and next_shift_type == "DAY":
        if rest < business_rules_for_given_shift_length['time_between_eve_shift_to_day_shift']:
            sufficient_rest = False
    elif prev_shift_type == "MID" and next_shift_type == "MID":
        if rest < business_rules_for_given_shift_length['time_between_mid_shift_to_any_shift']:
            sufficient_rest = False
    else:
        if rest < 8:
            sufficient_rest = False

    return sufficient_rest


# Check if transition between two shifts is "desirable"
def desirable_move_between_shifts(prev_shift, next_shift, shift_length):

    if isinstance(prev_shift, list) or isinstance(next_shift, list):
        return True

    if isinstance(prev_shift, str):
        pass
    else:
        prev_shift_type = determine_type_of_shift(prev_shift, shift_length)

    if isinstance(next_shift, str):
        pass
    else:
        next_shift_type = determine_type_of_shift(next_shift, shift_length)

    if next_shift == "X":
        return True

    if prev_shift == "X":
        return True

    if prev_shift_type == "EVE":
        if next_shift_type == "EVE" or next_shift_type == "DAY":
            return True
        else:
            return False
    elif prev_shift_type == "DAY":
        if next_shift_type == "DAY" or next_shift_type == "MID":
            return True
        else:
            return False
    else:
        if shift_length == 10:
            return True
        else:
            if next_shift_type == "MID":
                return True
            else:
                return False


def valid_number_of_consecutive_shifts(row, df, shift_type_to_check, shift_length, shift_to_add=False, add_shift_day=None):
    global business_rules
    business_rules_for_shift_length = business_rules[shift_length]
    max_number_of_consecutive_mid_shifts = business_rules_for_shift_length['number_of_consecutive_mid_shifts']

    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    # create dict representing shift line that need to be checked for consecutive mid-shifts
    dict_of_shift_line = {}
    for day in days:
        dict_of_shift_line[day] = df.at[row, day]

    # if we are testing if a shift can be added to a shift line, add this shift line before testing
    if shift_to_add == 0:
        dict_of_shift_line[add_shift_day] = shift_to_add
    elif shift_to_add:
        dict_of_shift_line[add_shift_day] = shift_to_add

    # look at every day
    for day in days:
        consecutive_shift_count = 0
        day_to_check = day
        shift = dict_of_shift_line[day_to_check]

        # track if we are dealing with an exception case (2100)
        exception_case = True

        # if we encounter a shift of type, keep counting next day until shift of type is not encountered
        while determine_type_of_shift(shift, shift_length) == shift_type_to_check:
            # during while loop, as soon as one non-exception shift is encounter, exception case no longer applies
            if shift != 21:
                exception_case = False
            consecutive_shift_count += 1
            day_to_check = next(day_to_check)
            shift = dict_of_shift_line[day_to_check]
            # check if the number of consecutive shift types exceeds allowed maximum without being exception case
            if consecutive_shift_count > max_number_of_consecutive_mid_shifts and not exception_case:
                return False

    return True


# If a cell with list of potential shifts == 0, this function attempts to find a shift already assigned on the same day
# Then attempts to fill in the impossible cell with this value and assign a new value to replace the given cell
def fill_impossible_cell(imp_row, day, df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    # Look at each cell in the column of the impossible cell
    for i in range(0, number_of_workers):
        # determine if cell already has an assigned shift
        if isinstance(df.at[i, day], int):
            # store the value of the shift
            given_shift = df.at[i, day]
            # determine if shift in cell could fill the impossible cell
            if sufficient_rest_between_shifts(df.at[imp_row, previous(day)], given_shift, shift_length) and sufficient_rest_between_shifts(given_shift, df.at[imp_row, next(day)], shift_length) and desirable_move_between_shifts(df.at[imp_row, previous(day)], given_shift, shift_length) and desirable_move_between_shifts(given_shift, df.at[imp_row, next(day)], shift_length) and valid_number_of_consecutive_shifts(imp_row, df, 'MID', shift_length, given_shift, day):
                # determine if there is an unassigned shift that could replace the given shift
                for shift, quantity in shifts_assigned_on[day].items():
                    if quantity < daily_shifts[day][shift]:
                        if sufficient_rest_between_shifts(df.at[i, previous(day)], shift, shift_length) and sufficient_rest_between_shifts(shift, df.at[i, next(day)], shift_length) and desirable_move_between_shifts(df.at[i, previous(day)], shift, shift_length) and desirable_move_between_shifts(shift, df.at[i, next(day)], shift_length) and valid_number_of_consecutive_shifts(i, df, 'MID', shift_length, shift, day):
                            # reset lists of potential shifts for all cells in row
                            for day_to_reset_potentials in days:
                                if day_to_reset_potentials != day:
                                    if isinstance(df.at[i, day_to_reset_potentials], list):
                                        df.at[i, day_to_reset_potentials] = list_of_potential_shifts(day_to_reset_potentials, shifts_assigned_on, daily_shifts)
                            # run cascade on all shifts already assigned
                            for day_to_cascade in days:
                                if day_to_cascade != day:
                                    if isinstance(df.at[i, day_to_cascade], int):
                                        try:
                                            cascade(i, day_to_cascade, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                                        except CascadeError:
                                            print('Cascade Error')
                            # assign new shift to replace given shift
                            new_shift_assigned = assign_shift(i, day, shift, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                            if new_shift_assigned:
                                shifts_assigned_on = new_shift_assigned
                                # take away count for shift I just removed
                                shifts_assigned_on[day][given_shift] -= 1

                                # reset lists of potential shifts for all cells in row
                                for day_to_reset_potentials in days:
                                    if day_to_reset_potentials != day:
                                        if isinstance(df.at[imp_row, day_to_reset_potentials], list):
                                            df.at[imp_row, day_to_reset_potentials] = list_of_potential_shifts(day_to_reset_potentials, shifts_assigned_on, daily_shifts)
                                # run cascade on all shifts already assigned
                                for day_to_cascade in days:
                                    if day_to_cascade != day:
                                        if isinstance(df.at[imp_row, day_to_cascade], int):
                                            try:
                                                cascade(imp_row, day_to_cascade, shifts_assigned_on, daily_shifts, df,shift_length, number_of_workers)
                                            except CascadeError:
                                                print('Cascade Error')

                                # assign given shift to impossible cell
                                given_shift_assigned = assign_shift(imp_row, day, given_shift, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                                if given_shift_assigned:
                                    shifts_assigned_on = given_shift_assigned
                                    return True
    return False


def get_potential_shift_to_swap(shift_to_swap, swap_row, day, df, shift_length, number_of_workers):
    # define shifts before and after shift to swap
    prev_shift_to_swap = df.at[swap_row, previous(day)]
    next_shift_to_swap = df.at[swap_row, next(day)]

    # Look at each row
    for i in range(0, number_of_workers):
        # Define key shifts
        prev_given_shift = df.at[i, previous(day)]
        given_shift = df.at[i, day]
        next_given_shift = df.at[i, next(day)]

        # skip any row where the given shift is RDO or unassigned:
        if given_shift == "X" or isinstance(given_shift, list):
            continue

        # Check if given shift can replace shift to swap
        given_can_replace_shift_to_swap = False
        shift_to_swap_can_replace_given_shift = False

        if sufficient_rest_between_shifts(prev_shift_to_swap, given_shift, shift_length):
            if desirable_move_between_shifts(prev_shift_to_swap, given_shift, shift_length):
                if sufficient_rest_between_shifts(given_shift, next_shift_to_swap, shift_length):
                    if desirable_move_between_shifts(given_shift, next_shift_to_swap, shift_length):
                        if valid_number_of_consecutive_shifts(swap_row, df, 'MID', shift_length, given_shift, day):
                            given_can_replace_shift_to_swap = True

        if given_can_replace_shift_to_swap:

            # Check if shift to swap can replace given shift
            if sufficient_rest_between_shifts(prev_given_shift, shift_to_swap, shift_length):
                if desirable_move_between_shifts(prev_given_shift, shift_to_swap, shift_length):
                    if sufficient_rest_between_shifts(shift_to_swap, next_given_shift, shift_length):
                        if desirable_move_between_shifts(shift_to_swap, next_given_shift, shift_length):
                            if valid_number_of_consecutive_shifts(i, df, 'MID', shift_length, shift_to_swap, day):
                                shift_to_swap_can_replace_given_shift = True
                                break

    if shift_to_swap_can_replace_given_shift:
        # Return the shift that can be swapped and its row
        return given_shift, i
    else:
        return False


# swap a series of shifts before or after an impossible shift
def swap_series_of_days(imp_row, imp_day, impossible_shift, direction, df, number_of_workers, shift_length, number_of_attempts, max_number_of_attempts):

    successful_swap = False

    if direction == "previous":
        first_series_day = previous(imp_day)
        last_series_day = previous(imp_day)
        before_first_imp_row_shift = df.at[imp_row, previous(first_series_day)]
        first_imp_row_shift = df.at[imp_row, first_series_day]
        last_imp_row_shift = df.at[imp_row, last_series_day]
        after_last_imp_row_shift = impossible_shift
    else:
        first_series_day = next(imp_day)
        last_series_day = next(imp_day)
        before_first_imp_row_shift = impossible_shift
        first_imp_row_shift = df.at[imp_row, first_series_day]
        last_imp_row_shift = df.at[imp_row, last_series_day]
        after_last_imp_row_shift = df.at[imp_row, next(last_series_day)]

    # create a dict that tracks if a row should be checked -- if a row has a series that contains an RDO mark to skip
    rows_to_check = {}
    for i in range(0, number_of_workers):
        rows_to_check[i] = True

    while not successful_swap:
        # Look at all rows
        for row, check_row in rows_to_check.items():
            # only look at rows that do not contain RDO in series
            if check_row == True:

                # define key given shifts
                before_first_given_row_shift = df.at[row, previous(first_series_day)]
                first_given_row_shift = df.at[row, first_series_day]
                if first_given_row_shift == "X" or isinstance(first_given_row_shift, list):
                    rows_to_check[row] = False
                    continue
                last_given_row_shift = df.at[row, last_series_day]
                if last_given_row_shift == "X" or isinstance(last_given_row_shift, list):
                    rows_to_check[row] = False
                    continue
                after_last_given_row_shift = df.at[row, next(last_series_day)]

                imp_series_can_replace_given_series = False
                given_series_can_replace_imp_series = False

                # determine if given series can replace impossible series
                if sufficient_rest_between_shifts(before_first_imp_row_shift, first_given_row_shift, shift_length):
                    if sufficient_rest_between_shifts(last_given_row_shift, after_last_imp_row_shift, shift_length):
                        if number_of_attempts < max_number_of_attempts - 2:
                            if desirable_move_between_shifts(before_first_imp_row_shift, first_given_row_shift, shift_length):
                                if desirable_move_between_shifts(last_given_row_shift, after_last_imp_row_shift, shift_length):
                                    given_series_can_replace_imp_series = True
                        else:
                            given_series_can_replace_imp_series = True

                if given_series_can_replace_imp_series:

                    # determine if impossible series can replace given series
                    if sufficient_rest_between_shifts(before_first_given_row_shift, first_imp_row_shift, shift_length):
                        if sufficient_rest_between_shifts(last_imp_row_shift, after_last_given_row_shift, shift_length):
                            if number_of_attempts < max_number_of_attempts - 2:
                                if desirable_move_between_shifts(before_first_given_row_shift, first_imp_row_shift, shift_length):
                                    if desirable_move_between_shifts(last_imp_row_shift, after_last_given_row_shift, shift_length):
                                        imp_series_can_replace_given_series = True
                            else:
                                imp_series_can_replace_given_series = True

                if imp_series_can_replace_given_series:
                    # check if number of consecutive mids violated by swap

                    # make a copy of impossible row
                    imp_row_df = df[df.index==imp_row]
                    imp_row_copy = imp_row_df.copy()

                    # make a copy of given row
                    given_row_df = df[df.index==row]
                    given_row_copy = given_row_df.copy()

                    # track if last shift in series has been swapped
                    last_shift_not_swapped = True

                    # Swap shifts on each day of the series into the copies of the rows
                    current_day_to_swap_shifts = first_series_day

                    while last_shift_not_swapped:
                        given_shift_to_swap = df.at[row, current_day_to_swap_shifts]
                        impossible_shift_to_swap = df.at[imp_row, current_day_to_swap_shifts]

                        given_row_copy[current_day_to_swap_shifts] = impossible_shift_to_swap
                        imp_row_copy[current_day_to_swap_shifts] = given_shift_to_swap

                        if current_day_to_swap_shifts == last_series_day:
                            last_shift_not_swapped = False
                        else:
                            current_day_to_swap_shifts = next(current_day_to_swap_shifts)

                    # check if copies of rows satisfy valid number of consecutive shifts
                    if valid_number_of_consecutive_shifts(row, given_row_copy, 'MID', shift_length) and valid_number_of_consecutive_shifts(imp_row, imp_row_copy, 'MID', shift_length):
                        successful_swap = True
                        # if so, assign row copies to dataframe
                        given_row_copy = given_row_copy.squeeze('rows')
                        imp_row_copy = imp_row_copy.squeeze('rows')
                        df.iloc[row] = given_row_copy
                        df.iloc[imp_row] = imp_row_copy

                    print('swapped values from given row', row, 'with impossible row', imp_row)
                    break

        if not successful_swap:
            # if all rows checked without solution, make series one larger
            if direction == "previous":
                first_series_day = previous(first_series_day)
                before_first_imp_row_shift = df.at[imp_row, previous(first_series_day)]
                first_imp_row_shift = df.at[imp_row, first_series_day]
                if first_imp_row_shift == "X" or isinstance(first_imp_row_shift, list):
                    break
            else:
                last_series_day = next(last_series_day)
                last_imp_row_shift = df.at[imp_row, last_series_day]
                if last_imp_row_shift == "X" or isinstance(last_imp_row_shift, list):
                    break
                after_last_imp_row_shift = df.at[imp_row, next(last_series_day)]

    if successful_swap:
        return True
    else:
        print("Swap NOT successful")
        return False


##### This is still in development / not being used
# Sometimes a shift when assigned can cause an impossible shift by cascade
# This attempts to swap a shift that caused an impossible on the previous cascade
# The goal is to swap it with a shift that occurs LATER allowing the previous cascade to not cause impossibility
def swap_later_day_causing_previous_impossible(row, impossible_day, df, shifts_assigned_on, daily_shifts, shift_length, number_of_workers):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    # identify problem causing day
    problem_causing_day = next(impossible_day)
    while not isinstance(df.at[row, problem_causing_day], int):
        problem_causing_day = next(problem_causing_day)

    # identify problem causing shift
    problem_causing_shift = df.at[row, problem_causing_day]

    # Create a list of shifts later than the problem causing shift
    shift_list = []
    for shift in daily_shifts['SUN'].keys():
        shift_list.append(shift)

    index_of_problem_causing_shift = shift_list.index(problem_causing_shift)
    shifts_later_than_problem_causing_shift = shift_list[index_of_problem_causing_shift + 1:]

    # Look at every potential later shift
    for later_shift in shifts_later_than_problem_causing_shift:

        # Reset the potential lists to the state before the problem shift was assigned
        for day_to_reset_potentials in days:
            if isinstance(df.at[row, day_to_reset_potentials], list):
                df.at[row, day_to_reset_potentials] = list_of_potential_shifts(day_to_reset_potentials, shifts_assigned_on, daily_shifts)
        # run cascade on all shifts already assigned (except problem day)
        for day_to_cascade in days:
            if day_to_cascade != problem_causing_day:
                if isinstance(df.at[row, day_to_cascade], int):
                    try:
                        cascade(row, day_to_cascade, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    except CascadeError:
                        print('Cascade Error')

        # track if this later shift now longer makes impossible day impossible
        later_shift_is_plausible = False

        # Test to see if sufficient rest between any potential shift from impossible day and the now later shift
        for potential_shift in df.at[row, impossible_day]:
            print(potential_shift)
            if sufficient_rest_between_shifts(potential_shift, later_shift, shift_length):
                later_shift_is_plausible = True
                break

        # If the later shift is plausible, try to find another row where this later shift has already been assigned
        if later_shift_is_plausible:
            for i in range(0, number_of_workers):
                given_shift = df.at[i, problem_causing_day]
                if given_shift == later_shift:
                    # See if problem shift can fit in given shift location
                    if sufficient_rest_between_shifts(df.at[i, previous(problem_causing_day)], problem_causing_shift, shift_length) and sufficient_rest_between_shifts(problem_causing_shift, df.at[i, next(problem_causing_day)], shift_length):
                        # See if given shift can fit in problem shift location
                        if sufficient_rest_between_shifts(df.at[row, previous(problem_causing_day)], given_shift, shift_length) and sufficient_rest_between_shifts(given_shift, df.at[row, next(problem_causing_day)], shift_length):
                            print('The problem shift', problem_causing_shift, 'can be swapped with the shift', given_shift)


# Attempt to assign shift based on rest restrictions
def assign_shift(row, day, shift, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers):
    # check if shift meets requirements to be assigned (assume not)
    shift_should_be_assigned = False
    return_value = False
    # check if current cell is list of potential shifts
    if shifts_assigned_on[day][shift] < daily_shifts[day][shift]:
        # check if previous shift is list/unassigned or an RDO
        if isinstance(df.at[row, previous(day)], list) or df.at[row, previous(day)] == 'X':
            shift_should_be_assigned = True
        # check if previous shift is a mid -> (only occurs for 10 hour schedules)
        elif determine_type_of_shift(df.at[row, previous(day)], shift_length) == "MID":
            shift_should_be_assigned = True
        # check if there is sufficient rest between the shift on the prior day
        # check if there is sufficient rest between the shift on the next day
        elif sufficient_rest_between_shifts(df.at[row, previous(day)], shift, shift_length) and sufficient_rest_between_shifts(shift, df.at[row, next(day)], shift_length) and desirable_move_between_shifts(df.at[row, previous(day)], shift, shift_length) and desirable_move_between_shifts(shift, df.at[row, next(day)], shift_length) and valid_number_of_consecutive_shifts(row, df, 'MID', shift_length, shift, day):
            shift_should_be_assigned = True

        if shift_should_be_assigned:
            # assign shift if conditions are met
            df.at[row, day] = shift
            # add assigned shift to totals assigned
            shifts_assigned_on[day][shift] += 1
            try:
                cascade(row, day, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
            except CascadeError:
                print('Cascade Error')
            return_value = shifts_assigned_on

            # Check if assigning this shift filled the requirements for the day
            if shifts_assigned_on[day][shift] == daily_shifts[day][shift]:
                for j in range(0, number_of_workers):
                    if isinstance(df.at[j, day], list):
                        if shift in df.at[j, day]:
                            df.at[j, day].remove(shift)
                            if len(df.at[j, day]) == 1:
                                assigned_shift = assign_shift(j, day, df.at[j, day][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                                if assigned_shift:
                                    shifts_assigned_on = assigned_shift
                                    return_value = shifts_assigned_on
                                try:
                                    cascade(j, day, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                                except CascadeError:
                                    print('Cascade Error')
                            elif len(df.at[j, day]) == 0:
                                fill_impossible_cell(j, day, df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)
    else:
        if isinstance(df.at[row, day], list):
            try:
                df.at[row, day].remove(shift)
            except ValueError:
                pass

    return return_value


def cascade(row, day, shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers):

    if isinstance(df.at[row, day], list):
        current_day_shifts = df.at[row, day].copy()
    else:
        current_day_shifts = df.at[row, day]

    ### cascade in previous direction
    previous_day_affected = False
    previous_cascade_complete = False

    # Look at previous column/day
    if isinstance(df.at[row, previous(day)], list):
        previous_day_shifts = df.at[row, previous(day)].copy()
    else:
        previous_day_shifts = df.at[row, previous(day)]

    # case 1: RDO - stop process
    if previous_day_shifts == 'X':
        previous_cascade_complete = True
    # case 2: Integer/Float
    elif isinstance(previous_day_shifts, int) or isinstance(previous_day_shifts, float):
        # sub_case 1: Integer/Float
        if isinstance(current_day_shifts, int) or isinstance(current_day_shifts, float):
            if sufficient_rest_between_shifts(previous_day_shifts, current_day_shifts, shift_length) and desirable_move_between_shifts(previous_day_shifts, current_day_shifts, shift_length):
                previous_cascade_complete = True
            else:
                raise CascadeError
        # sub_case 2: list
        elif isinstance(current_day_shifts, list):
            for current_day_shift in current_day_shifts:
                if not sufficient_rest_between_shifts(previous_day_shifts, current_day_shift, shift_length) or not desirable_move_between_shifts(previous_day_shifts, current_day_shift, shift_length):
                    df.at[row, day].remove(current_day_shift)
            if len(df.at[row, day]) > 1:
                previous_cascade_complete = True
            elif len(df.at[row, day]) == 1:
                assigned_shift = assign_shift(row, day, df.at[row, day][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                if assigned_shift:
                    shifts_assigned_on = assigned_shift
                previous_cascade_complete = True
            else:
                fill_impossible_cell(row, day, df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)
    # case 3: List
    elif isinstance(previous_day_shifts, list):
        # sub_case 1: Integer/Float
        if isinstance(current_day_shifts, int) or isinstance(current_day_shifts, float):
            for previous_day_shift in previous_day_shifts:
                if not sufficient_rest_between_shifts(previous_day_shift, current_day_shifts, shift_length) or not desirable_move_between_shifts(previous_day_shift, current_day_shifts, shift_length):
                    df.at[row, previous(day)].remove(previous_day_shift)
                    previous_day_affected = True
            if previous_day_affected:
                if len(df.at[row, previous(day)]) > 1:
                    try:
                        cascade(row, previous(day), shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    except CascadeError:
                        print('Cascade Error')
                    previous_cascade_complete = True
                elif len(df.at[row, previous(day)]) == 1:
                    assigned_shift = assign_shift(row, previous(day), df.at[row, previous(day)][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    if assigned_shift:
                        shifts_assigned_on = assigned_shift
                    previous_cascade_complete = True
                else:
                    fill_impossible_cell(row, previous(day), df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)

            else:
                previous_cascade_complete = True
        # sub_case 2: list
        elif isinstance(current_day_shifts, list):
            for previous_day_shift in previous_day_shifts:
                previous_shift_should_be_removed = True
                for current_day_shift in current_day_shifts:
                    if sufficient_rest_between_shifts(previous_day_shift, current_day_shift, shift_length) and desirable_move_between_shifts(previous_day_shift, current_day_shift, shift_length):
                        previous_shift_should_be_removed = False
                        break
                if previous_shift_should_be_removed:
                    df.at[row, previous(day)].remove(previous_day_shift)
                    previous_day_affected = True
            if previous_day_affected:
                if len(df.at[row, previous(day)]) > 1:
                    try:
                        cascade(row, previous(day), shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    except CascadeError:
                        print('Cascade Error')
                    previous_cascade_complete = True
                elif len(df.at[row, previous(day)]) == 1:
                    assigned_shift = shift_assigned = assign_shift(row, previous(day), df.at[row, previous(day)][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    if assigned_shift:
                        shifts_assigned_on = assigned_shift
                    previous_cascade_complete = True
                else:
                    fill_impossible_cell(row, previous(day), df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)
            else:
                previous_cascade_complete = True

    ###cascade in forward direction

    if isinstance(df.at[row, day], list):
        current_day_shifts = df.at[row, day].copy()
    else:
        current_day_shifts = df.at[row, day]

    next_day_affected = False
    next_cascade_complete = False

    # Look at next column/day
    if isinstance(df.at[row, next(day)], list):
        next_day_shifts = df.at[row, next(day)].copy()
    else:
        next_day_shifts = df.at[row, next(day)]

    # case 1: RDO - stop process
    if next_day_shifts == 'X':
        next_cascade_complete = True
    # case 2: Integer/Float
    elif isinstance(next_day_shifts, int) or isinstance(next_day_shifts, float):
        # sub_case 1: Integer/Float
        if isinstance(current_day_shifts, int) or isinstance(current_day_shifts, float):
            if sufficient_rest_between_shifts(current_day_shifts, next_day_shifts, shift_length) and desirable_move_between_shifts(current_day_shifts, next_day_shifts, shift_length):
                next_cascade_complete = True
            else:
                raise CascadeError
        # sub_case 2: list
        elif isinstance(current_day_shifts, list):
            for current_day_shift in current_day_shifts:
                if not sufficient_rest_between_shifts(current_day_shift, next_day_shifts, shift_length) or not desirable_move_between_shifts(current_day_shift, next_day_shifts, shift_length):
                    df.at[row, day].remove(current_day_shift)
            if len(df.at[row, day]) > 1:
                next_cascade_complete = True
            elif len(df.at[row, day]) == 1:
                assigned_shift = shift_assigned = assign_shift(row, day, df.at[row, day][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                shifts_assigned_on = assigned_shift
                next_cascade_complete = True
            else:
                fill_impossible_cell(row, day, df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)
    # case 3: List
    elif isinstance(next_day_shifts, list):
        # sub_case 1: Integer/Float
        if isinstance(current_day_shifts, int) or isinstance(current_day_shifts, float):
            for next_day_shift in next_day_shifts:
                if not sufficient_rest_between_shifts(current_day_shifts, next_day_shift, shift_length) or not desirable_move_between_shifts(current_day_shifts, next_day_shift, shift_length):
                    df.at[row, next(day)].remove(next_day_shift)
                    next_day_affected = True
            if next_day_affected:
                if len(df.at[row, next(day)]) > 1:
                    try:
                        cascade(row, next(day), shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    except CascadeError:
                        print('Cascade Error')
                    next_cascade_complete = True
                elif len(df.at[row, next(day)]) == 1:
                    assigned_shift = shift_assigned = assign_shift(row, next(day), df.at[row, next(day)][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    if assigned_shift:
                        shifts_assigned_on = assigned_shift
                    next_cascade_complete = True
                else:
                    fill_impossible_cell(row, next(day), df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)
            else:
                next_cascade_complete = True
        # sub_case 2: list
        elif isinstance(current_day_shifts, list):
            for next_day_shift in next_day_shifts:
                next_shift_should_be_removed = True
                for current_day_shift in current_day_shifts:
                    if sufficient_rest_between_shifts(current_day_shift, next_day_shift, shift_length) and desirable_move_between_shifts(current_day_shift, next_day_shift, shift_length):
                        next_shift_should_be_removed = False
                        break
                if next_shift_should_be_removed:
                    df.at[row, next(day)].remove(next_day_shift)
                    next_day_affected = True
            if next_day_affected:
                if len(df.at[row, next(day)]) > 1:
                    try:
                        cascade(row, next(day), shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    except CascadeError:
                        print('Cascade Error')
                    next_cascade_complete = True
                elif len(df.at[row, next(day)]) == 1:
                    assigned_shift = assign_shift(row, next(day), df.at[row, next(day)][0], shifts_assigned_on, daily_shifts, df, shift_length, number_of_workers)
                    if assigned_shift:
                        shifts_assigned_on = assigned_shift
                    next_cascade_complete = True
                else:
                    fill_impossible_cell(row, next(day), df, shifts_assigned_on, daily_shifts, number_of_workers, shift_length)
            else:
                next_cascade_complete = True


def det_impossible_cells_to_fill(df, number_of_workers):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    impossible_cell_exists = False
    list_of_impossible_cells = {}

    # Look at each row
    for i in range(0, number_of_workers):
        # Look at each day
        for day in days:
            # Check if any cell is impossible to fill
            if isinstance(df.at[i, day], list):
                if len(df.at[i, day]) == 0:
                    try:
                        list_of_impossible_cells[day].append(i)
                    except KeyError:
                        list_of_impossible_cells[day] = [i]
                    impossible_cell_exists = True

    if impossible_cell_exists:
        return list_of_impossible_cells
    else:
        return False


def determine_number_of_each_shift(df, number_of_workers):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    shift_totals = {}

    for day in days:
        shift_totals[day] = {}
        shifts = []

        for i in range(0, number_of_workers):
            shift = df.at[i, day]
            if shift == 'X' or isinstance(shift, list):
                continue
            else:
                if shift not in shifts:
                    shifts.append(shift)
                    shift_totals[day][shift] = 0
                shift_totals[day][shift] += 1

    return shift_totals


def determine_number_of_each_type_of_shift(df, number_of_workers, shift_length):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    shift_type_totals = {}

    for day in days:
        shift_type_totals[day] = {}
        shift_type_totals[day]['EVE'] = 0
        shift_type_totals[day]['DAY'] = 0
        shift_type_totals[day]['MID'] = 0

        for i in range(0, number_of_workers):
            if df.at[i, day] == 'X' or isinstance(df.at[i, day], list):
                continue
            else:
                shift_type = determine_type_of_shift(df.at[i, day], shift_length)
                if shift_type == 'EVE':
                    shift_type_totals[day]['EVE'] += 1
                elif shift_type == 'DAY':
                    shift_type_totals[day]['DAY'] += 1
                else:
                    shift_type_totals[day]['MID'] += 1

    return shift_type_totals


def clear_empty_days(missing_shifts):

    cleared_missing_shifts = {}

    for day in missing_shifts.keys():
        if len(missing_shifts[day]) == 0:
            continue
        else:
            cleared_missing_shifts[day] = {}
            for shift, quantity in missing_shifts[day].items():
                cleared_missing_shifts[day][shift] = quantity

    return cleared_missing_shifts


def determine_number_of_missing_shifts_per_day(number_of_each_shift, daily_shifts):

    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    missing_shifts_on = {}

    for day in days:
        for shift in daily_shifts[day].keys():
            try:
                num_shifts_assigned = number_of_each_shift[day][shift]
            except KeyError:
                num_shifts_assigned = 0
            num_shifts_required = daily_shifts[day][shift]
            num_of_missing_shifts = num_shifts_required - num_shifts_assigned
            if num_of_missing_shifts > 0:
                try:
                    missing_shifts_on[day][shift] = num_of_missing_shifts
                except KeyError:
                    missing_shifts_on[day] = {}
                    missing_shifts_on[day][shift] = num_of_missing_shifts

    return missing_shifts_on


# This function is not currently being used
# swap out shifts in successful schedule that are causing undesirable moves
def swap_undesirable_move_shift(row, day, df, shift_length, number_of_workers):
    previous_undesirable_shift = df.at[row, previous(day)]
    undesirable_shift = df.at[row, day]
    following_undesirable_shift = df.at[row, next(day)]

    # Look at every shift on the same day
    for i in range(0, number_of_workers):
        if i != row:
            # identify the previous, given, and next shift on that day
            given_shift = df.at[i, day]
            if given_shift == 'X':
                continue
            previous_given_shift = df.at[i, previous(day)]
            following_given_shift = df.at[i, next(day)]

            # see if given_shift would satisfy rest rules with shifts before and after
            if sufficient_rest_between_shifts(previous_undesirable_shift, given_shift, shift_length) and sufficient_rest_between_shifts(given_shift, following_undesirable_shift, shift_length):
                # see if given_shift would satisfy desirable moves
                if desirable_move_between_shifts(previous_undesirable_shift, given_shift, shift_length) and desirable_move_between_shifts(given_shift, following_undesirable_shift, shift_length):
                    # see if undesirable shift would satisfy rest rules
                    if sufficient_rest_between_shifts(previous_given_shift, undesirable_shift, shift_length) and sufficient_rest_between_shifts(undesirable_shift, following_given_shift, shift_length):
                        # see if undesirable shift would satisfy desirable moves:
                        if desirable_move_between_shifts(previous_given_shift, undesirable_shift, shift_length) and desirable_move_between_shifts(undesirable_shift, following_given_shift, shift_length):
                            # if all conditions met, swap given shifts
                            df.at[i, day] = undesirable_shift
                            df.at[row, day] = given_shift
                            return True

    # if entire loop executes, swap failed
    return False


# returns the following day
def next(day):
    if day == "SUN":
        return "MON"
    elif day == "MON":
        return "TUE"
    elif day == "TUE":
        return "WED"
    elif day == "WED":
        return "THU"
    elif day == "THU":
        return "FRI"
    elif day == "FRI":
        return "SAT"
    elif day == "SAT":
        return "SUN"
    else:
        return "invalid entry"


# returns the previous day
def previous(day):
    if day == "SUN":
        return "SAT"
    elif day == "MON":
        return "SUN"
    elif day == "TUE":
        return "MON"
    elif day == "WED":
        return "TUE"
    elif day == "THU":
        return "WED"
    elif day == "FRI":
        return "THU"
    elif day == "SAT":
        return "FRI"
    else:
        return "invalid entry"


def dataframe_int_to_mt(df):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    for row in range(0, len(df)):
        for day in days:
            shift_int = df.at[row, day]
            if shift_int == 'X':
                df.at[row, day] = shift_int
            elif shift_int <= 9:
                df.at[row, day] = "0" + str(shift_int) + ":00"
            elif shift_int >= 10:
                df.at[row, day] = str(shift_int) + ":00"
    return df


def dataframe_int_to_st(df):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    for row in range(0, len(df)):
        for day in days:
            if isinstance(df.at[row, day], list):
                continue
            elif isinstance(df.at[row, day], float):
                decimal_value = df.at[row, day]
                hours = str(decimal_value).split('.')[0]
                mins = int(round((decimal_value % 1) * 60, 0))
                if mins < 10:
                    mins = str(0) + str(mins)
                else:
                    mins = str(mins)

                df.at[row, day] = hours + mins

            else:
                shift_int = df.at[row, day]
                if shift_int == 'X':
                    df.at[row, day] = shift_int
                elif shift_int == 0:
                    df.at[row, day] = 'N'
                elif shift_int == 23:
                    df.at[row, day] = 'M'
                else:
                    if df.at[row, day] in [6, 7, 8, 13, 15, 16]:
                        df.at[row, day] = shift_int % 12
                    else:
                        df.at[row, day] = str(shift_int) + "00"
    return df


def create_shift_totals_df(total_shifts, daily_shifts):
    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    # create dataframe to hold number of each shift assigned per day
    total_shifts_df = pd.DataFrame(columns=days)

    for shift in list(daily_shifts['SUN'].keys()):
        df_to_append = pd.DataFrame([[0, 0, 0, 0, 0, 0, 0]], columns=days, index=[shift])
        total_shifts_df = total_shifts_df.append(df_to_append)

    for day in days:
        for shift in total_shifts[day].keys():
            total_shifts_df.at[shift, day] = total_shifts[day][shift]

    return total_shifts_df


def get_missing_shift_count(missing_shifts):
    missing_shift_count = 0

    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    for day in days:
        try:
            for shift, quantity in missing_shifts[day].items():
                missing_shift_count += quantity
        except KeyError:
            continue

    return missing_shift_count


def fill_in_missing_shifts(df, total_shifts, daily_shifts, number_of_workers, shift_length):

    days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    missing_shifts_filled = True

    for day in days:
        for i in range(0, number_of_workers):
            if isinstance(df.at[i, day], list):
                for shift in daily_shifts[day].keys():
                    if daily_shifts[day][shift] != 0:
                        if sufficient_rest_between_shifts(df.at[i, previous(day)], shift, shift_length):
                            if desirable_move_between_shifts(df.at[i, previous(day)], shift, shift_length):
                                if sufficient_rest_between_shifts(shift, df.at[i, next(day)], shift_length):
                                    if desirable_move_between_shifts(shift, df.at[i, next(day)], shift_length):
                                        df.at[i, day] = shift
                                        try:
                                            total_shifts[day][shift] += 1
                                        except KeyError:
                                            total_shifts[day][shift] = 1
                                        break
                if isinstance(df.at[i, day], list):
                    missing_shifts_filled = False

    if missing_shifts_filled:
        missing_shifts_remaining = False
        return df, total_shifts, missing_shifts_remaining
    else:
        print('Some missing shifts could not be filled')
        missing_shifts_remaining = True
        return df, total_shifts, missing_shifts_remaining


### This function is not currently being used AND is incomplete
### Allows adding of addition row to finish schedule with one missing shift
def add_missing_shift(missing_shifts, daily_shifts, df, shift_length):

    days = {"SUN": 0, "MON": 1, "TUE": 2, "WED": 3, "THU": 4, "FRI": 5, "SAT": 6}

    potential_shifts_to_add = {}
    for day in days:
        potential_shifts_to_add[day] = []
        for shift in daily_shifts[day].keys():
            if 22 <= shift < 24 or 0 <= shift < 6:
                continue
            elif daily_shifts[day][shift] == 0:
                continue
            else:
                potential_shifts_to_add[day].append(shift)

    for day in days:
        try:
            for shift, quantity in missing_shifts.items():
                # determine type of shift added -- assign RDO accordingly
                type_of_shift_to_assign = determine_type_of_shift(shift, shift_length)
                while quantity > 0:
                    # create blank list and assign missing shift to the required day
                    row_to_add = [None] * 7
                    row_to_add[days[day]] = shift

                    # Assign RDO based on type of shift
                    if type_of_shift_to_assign == 'EVE':
                        row_to_add[day[previous(day)]] = 'X'
                        row_to_add[day[previous(previous(day))]] = 'X'
                    elif type_of_shift_to_assign == 'DAY' or type_of_shift_to_assign == 'MID':
                        row_to_add[day[next(day)]] = 'X'
                        row_to_add[day[next(next(day))]] = 'X'


                    ################## Assign remaining shifts (based on busiest shifts)

                    quantity -= 1
        except KeyError:
            continue


